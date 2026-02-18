#!/usr/bin/env python3

import argparse
import os
import statistics
from collections import defaultdict
from typing import Dict, List, Tuple, Optional

# -----------------------------------
# CLI
# -----------------------------------
parser = argparse.ArgumentParser(
    description="Step3 compile: BLAST top hits + variant-only site evidence across isolates + combined status tables."
)
parser.add_argument(
    "-o", "--outdir", required=True,
    help="Pipeline output root (contains per-isolate subdirs + summary/)"
)
parser.add_argument(
    "-l", "--linelist", required=True,
    help="Linelist (CSV or TSV). Isolate must be in first column. Run column is ignored."
)
parser.add_argument(
    "-f", "--targets-fasta", required=True,
    help="Targets FASTA used in Step 1 BLAST (e.g., scripts/nucleotide.fna)"
)
parser.add_argument(
    "-s", "--isolate", default="",
    help="Compile only one isolate (optional)"
)
parser.add_argument(
    "-r", "--ref-contig", default="CU458896.1",
    help="Reference contig name (only used for depth lookup)"
)
args = parser.parse_args()

OUTDIR = os.path.abspath(args.outdir)
LINELIST = os.path.abspath(args.linelist)
TARGETS_FASTA = os.path.abspath(args.targets_fasta)
ONLY_ISO = args.isolate.strip()
REF_CONTIG = args.ref_contig.strip()

os.makedirs(os.path.join(OUTDIR, "summary"), exist_ok=True)
os.makedirs(os.path.join(OUTDIR, "status"), exist_ok=True)

# -----------------------------------
# Target gene coordinates (MATCH 02_map_call.slurm)
# 1-based inclusive
# -----------------------------------
RRS_START, RRS_END = 1462398, 1463901
RRL_START, RRL_END = 1464208, 1467319
ERM41_START, ERM41_END = 2345955, 2346476

GENE_STARTS = {
    "rrs": RRS_START,
    "rrl": RRL_START,
    "erm41": ERM41_START,
}

# Sites-of-interest (gene coords, 1-based)
SITES: List[Tuple[str, int]] = [
    ("rrl", 2269), ("rrl", 2270), ("rrl", 2271), ("rrl", 2281), ("rrl", 2293),
    ("erm41", 19), ("erm41", 28),
    ("rrs", 1373), ("rrs", 1375), ("rrs", 1376), ("rrs", 1458),
]

# -----------------------------------
# erm41 truncation metrics (coverage-based; Criterion A prep)
# -----------------------------------
ERM41_LEFT_FLANK = (20, 140)     # gene positions (1-based, inclusive)
ERM41_RIGHT_FLANK = (450, 560)   # gene positions (1-based, inclusive)
ERM41_DEL_RANGE = (159, 432)     # gene positions (1-based, inclusive)
ERM41_CALLABLE_MIN_FLANK_MED = 20  # "callable" threshold on flank median depth

# -----------------------------------
# Helpers
# -----------------------------------
def read_isolates_from_linelist(path: str) -> List[str]:
    """
    Accept CSV or TSV. Header optional.
    Uses first column as isolate.
    """
    isolates: List[str] = []
    with open(path, "r") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            # split on tab OR comma (prefer tab if present)
            if "\t" in line:
                parts = line.split("\t")
            else:
                parts = line.split(",")
            iso = parts[0].strip()
            if iso.lower() in ("isolate", "#isolate"):
                continue
            isolates.append(iso)
    return isolates

def read_fasta_lengths(path: str) -> Dict[str, int]:
    lens: Dict[str, int] = {}
    name: Optional[str] = None
    seq_len = 0
    with open(path, "r") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            if line.startswith(">"):
                if name is not None:
                    lens[name] = seq_len
                name = line[1:].split()[0]
                seq_len = 0
            else:
                seq_len += len(line)
        if name is not None:
            lens[name] = seq_len
    return lens

def gene_group_from_query(qid: str) -> str:
    if qid.startswith("erm55"):
        return "erm55"
    if qid.startswith("erm39"):
        return "erm39"
    if qid.startswith("erm41"):
        return "erm41"
    return qid

def safe_float(x: str) -> float:
    try:
        return float(x)
    except Exception:
        return float("nan")

def safe_int(x: str) -> int:
    try:
        return int(x)
    except Exception:
        return -1

def compute_af(dp_str: str, ad_str: str) -> str:
    """
    AF = alt_count / DP, using AD (ref,alt) when available.
    Returns '' if cannot compute.
    """
    dp = safe_int(dp_str)
    if dp <= 0:
        dp = 0

    # AD usually like "ref,alt" (or sometimes "ref,alt,alt2")
    ad_parts = [p for p in ad_str.split(",") if p != ""]
    if len(ad_parts) < 2:
        return ""

    ref_ct = safe_int(ad_parts[0])
    alt_ct = safe_int(ad_parts[1])
    if alt_ct < 0:
        return ""

    denom = dp if dp > 0 else (ref_ct + alt_ct)
    if denom <= 0:
        return ""

    return f"{alt_ct / denom:.4f}"

def parse_blast_outfmt6(blast_tsv: str, qlens: Dict[str, int]) -> List[dict]:
    """
    step1 uses outfmt:
      qseqid sseqid pident length qlen slen qstart qend sstart send evalue bitscore qcovs
    """
    rows: List[dict] = []
    if not os.path.exists(blast_tsv) or os.path.getsize(blast_tsv) == 0:
        return rows

    with open(blast_tsv, "r") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split("\t")
            if len(parts) < 13:
                continue
            qseqid, sseqid = parts[0], parts[1]
            pident = safe_float(parts[2])
            alen = safe_int(parts[3])
            qlen = safe_int(parts[4])
            slen = safe_int(parts[5])
            qstart = safe_int(parts[6])
            qend = safe_int(parts[7])
            sstart = safe_int(parts[8])
            send = safe_int(parts[9])
            evalue = safe_float(parts[10])
            bitscore = safe_float(parts[11])
            qcovs = safe_float(parts[12])

            # If qlen in blast is 0/blank, fall back to targets fasta length
            qlen2 = qlen if qlen > 0 else qlens.get(qseqid, 0)

            rows.append({
                "query": qseqid,
                "group": gene_group_from_query(qseqid),
                "qlen": qlen2,
                "subject": sseqid,
                "pident": pident,
                "alen": alen,
                "qcov": qcovs,
                "slen": slen,
                "qstart": qstart,
                "qend": qend,
                "sstart": sstart,
                "send": send,
                "evalue": evalue,
                "bitscore": bitscore,
            })
    return rows

def choose_top_blast(hits: List[dict]) -> Tuple[Optional[dict], int, int]:
    """
    preferred: pident>=90 and qcov>=90
    rank: max(qcov), max(pident), max(bitscore), min(evalue)
    """
    if not hits:
        return None, 0, 0
    total = len(hits)
    preferred = [h for h in hits if (h["pident"] >= 90.0 and h["qcov"] >= 90.0)]
    pref_n = len(preferred)

    def keyfn(h):
        ev = h["evalue"]
        if ev != ev:  # nan
            ev = 1e300
        return (h["qcov"], h["pident"], h["bitscore"], -ev)

    pool = preferred if preferred else hits
    top = sorted(pool, key=keyfn, reverse=True)[0]
    return top, total, pref_n

def read_depth_tsv(depth_tsv: str, ref_contig: str) -> Dict[int, int]:
    """
    samtools depth output: CHROM POS DEPTH
    """
    depths: Dict[int, int] = {}
    if not os.path.exists(depth_tsv) or os.path.getsize(depth_tsv) == 0:
        return depths
    with open(depth_tsv, "r") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split("\t")
            if len(parts) < 3:
                continue
            chrom = parts[0]
            if chrom != ref_contig:
                continue
            pos = safe_int(parts[1])
            dep = safe_int(parts[2])
            if pos >= 1:
                depths[pos] = dep if dep >= 0 else 0
    return depths

def read_targets_variants_tsv(path: str) -> Dict[Tuple[str, int], dict]:
    """
    Parse Step2 output:
      Isolate Gene CHROM POS_ref POS_gene REF ALT QUAL DP AD

    Returns dict keyed by (gene, pos_gene).
    """
    d: Dict[Tuple[str, int], dict] = {}
    if not os.path.exists(path) or os.path.getsize(path) == 0:
        return d

    with open(path, "r") as f:
        _header = f.readline()
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split("\t")
            if len(parts) < 10:
                continue
            gene = parts[1]
            pos_gene = safe_int(parts[4])
            if pos_gene < 1:
                continue
            d[(gene, pos_gene)] = {
                "REF": parts[5],
                "ALT": parts[6],
                "QUAL": parts[7],
                "DP": parts[8],
                "AD": parts[9],
            }
    return d

def concat_status_files(outdir: str, isolates: List[str]) -> None:
    """
    Concatenate per-isolate status_step1.tsv and status_step2.tsv into:
      <outdir>/status/status_step1_all.tsv
      <outdir>/status/status_step2_all.tsv
    Keeps a single header row.
    """
    status_dir = os.path.join(outdir, "status")
    os.makedirs(status_dir, exist_ok=True)

    def _concat(filename: str, outname: str) -> None:
        outpath = os.path.join(status_dir, outname)
        header_written = False

        with open(outpath, "w") as out:
            for iso in isolates:
                p = os.path.join(outdir, iso, filename)
                if not os.path.exists(p) or os.path.getsize(p) == 0:
                    continue

                with open(p, "r") as f:
                    lines = f.readlines()

                if not lines:
                    continue

                # First line is header in your status files
                hdr = lines[0].rstrip("\n")
                if not header_written:
                    out.write(hdr + "\n")
                    header_written = True

                for line in lines[1:]:
                    line = line.rstrip("\n")
                    if line:
                        out.write(line + "\n")

            if not header_written:
                # No files found; still produce a consistent header
                out.write("Isolate\tStep\tStatus\tMessage\n")

    _concat("status_step1.tsv", "status_step1_all.tsv")
    _concat("status_step2.tsv", "status_step2_all.tsv")

def _median_or_blank(vals: List[int]) -> str:
    if not vals:
        return ""
    try:
        return f"{statistics.median(vals):.3f}"
    except Exception:
        return ""

# -----------------------------------
# Main
# -----------------------------------
isolates = read_isolates_from_linelist(LINELIST)
if ONLY_ISO:
    isolates = [x for x in isolates if x == ONLY_ISO]

qlens = read_fasta_lengths(TARGETS_FASTA)

# -----------------------
# BLAST summary
# -----------------------
blast_out = os.path.join(OUTDIR, "summary", "blast_top_hits.tsv")
blast_header = [
    "Isolate","GeneGroup","QueryID","QueryLen","Subject",
    "Pident","AlnLen","QcovPct",
    "Qstart","Qend","Sstart","Send","Evalue","Bitscore",
    "TotalHits","PreferredHits"
]

with open(blast_out, "w") as out:
    out.write("\t".join(blast_header) + "\n")

    for iso in isolates:
        blast_tsv = os.path.join(OUTDIR, iso, "blast", "blast_raw.tsv")
        hits = parse_blast_outfmt6(blast_tsv, qlens)

        by_group: Dict[str, List[dict]] = defaultdict(list)
        for h in hits:
            by_group[h["group"]].append(h)

        for group in ["erm41", "erm39", "erm55"]:
            top, total, pref_n = choose_top_blast(by_group.get(group, []))
            if top is None:
                out.write("\t".join([iso, group] + [""]*13 + ["0","0"]) + "\n")
                continue
            out.write("\t".join([
                iso,
                group,
                top["query"],
                str(top["qlen"]),
                top["subject"],
                f'{top["pident"]:.3f}',
                str(top["alen"]),
                f'{top["qcov"]:.3f}',
                str(top["qstart"]),
                str(top["qend"]),
                str(top["sstart"]),
                str(top["send"]),
                f'{top["evalue"]:.6g}' if top["evalue"] == top["evalue"] else "",
                f'{top["bitscore"]:.3f}' if top["bitscore"] == top["bitscore"] else "",
                str(total),
                str(pref_n),
            ]) + "\n")

# -----------------------
# Sites evidence (VARIANT-ONLY)
#   - Uses targets_variants.tsv for presence
#   - Uses depth for depth at the reference coordinate
#   - Output columns EXCLUDE POS_ref and VCF_present
# -----------------------
sites_out = os.path.join(OUTDIR, "summary", "sites_evidence.tsv")
sites_header = [
    "Isolate", "Gene", "position", "Depth",
    "REF", "ALT", "QUAL", "DP", "AD", "AF"
]

with open(sites_out, "w") as out:
    out.write("\t".join(sites_header) + "\n")

    for iso in isolates:
        depth_tsv = os.path.join(OUTDIR, iso, "variants", "targets_depth.tsv")
        var_tsv = os.path.join(OUTDIR, iso, "variants", "targets_variants.tsv")

        depths = read_depth_tsv(depth_tsv, REF_CONTIG)
        varmap = read_targets_variants_tsv(var_tsv)

        for gene, pos_gene in SITES:
            key = (gene, pos_gene)
            if key not in varmap:
                continue

            start = GENE_STARTS[gene]
            pos_ref = start + (pos_gene - 1)
            depth = depths.get(pos_ref, "")
            v = varmap[key]
            af = compute_af(v.get("DP",""), v.get("AD",""))

            out.write("\t".join([
                iso,
                gene,
                str(pos_gene),
                str(depth),
                v.get("REF",""),
                v.get("ALT",""),
                v.get("QUAL",""),
                v.get("DP",""),
                v.get("AD",""),
                af,
            ]) + "\n")

# -----------------------
# erm41 truncation metrics summary (coverage-based; interpretation later)
# -----------------------
trunc_out = os.path.join(OUTDIR, "summary", "erm41_truncation_metrics.tsv")
trunc_header = [
    "Isolate",
    "erm_len",
    "erm_median_depth",
    "left_median_depth",
    "right_median_depth",
    "flank_median_depth",
    "erm_del_range",
    "del_median_depth",
    "del_to_flank_ratio",
    "erm41_callable",
]

erm_len = ERM41_END - ERM41_START + 1

def _depth_list(depths: Dict[int, int], start_ref: int, end_ref: int) -> List[int]:
    # inclusive range
    return [depths.get(pos, 0) for pos in range(start_ref, end_ref + 1)]

def _gene_to_ref_range(gene_start_ref: int, g1: int, g2: int) -> Tuple[int, int]:
    # gene positions are 1-based inclusive
    r1 = gene_start_ref + (g1 - 1)
    r2 = gene_start_ref + (g2 - 1)
    return (r1, r2)

left_r1, left_r2 = _gene_to_ref_range(ERM41_START, ERM41_LEFT_FLANK[0], ERM41_LEFT_FLANK[1])
right_r1, right_r2 = _gene_to_ref_range(ERM41_START, ERM41_RIGHT_FLANK[0], ERM41_RIGHT_FLANK[1])
del_r1, del_r2 = _gene_to_ref_range(ERM41_START, ERM41_DEL_RANGE[0], ERM41_DEL_RANGE[1])
del_range_str = f"{ERM41_DEL_RANGE[0]}-{ERM41_DEL_RANGE[1]}"

with open(trunc_out, "w") as out:
    out.write("\t".join(trunc_header) + "\n")

    for iso in isolates:
        depth_tsv = os.path.join(OUTDIR, iso, "variants", "targets_depth.tsv")
        depths = read_depth_tsv(depth_tsv, REF_CONTIG)

        if not depths:
            out.write("\t".join([
                iso,
                str(erm_len),
                "", "", "", "",
                del_range_str,
                "", "",
                "FALSE",
            ]) + "\n")
            continue

        gene_depths = _depth_list(depths, ERM41_START, ERM41_END)
        left_depths = _depth_list(depths, left_r1, left_r2)
        right_depths = _depth_list(depths, right_r1, right_r2)
        del_depths = _depth_list(depths, del_r1, del_r2)

        erm_med = _median_or_blank(gene_depths)
        left_med = _median_or_blank(left_depths)
        right_med = _median_or_blank(right_depths)

        # flank median = median of the two flank medians (using numeric values)
        flank_med_val = None
        try:
            lm = statistics.median(left_depths) if left_depths else 0
            rm = statistics.median(right_depths) if right_depths else 0
            flank_med_val = statistics.median([lm, rm])
            flank_med = f"{flank_med_val:.3f}"
        except Exception:
            flank_med = ""
            flank_med_val = None

        del_med_val = None
        try:
            del_med_val = statistics.median(del_depths) if del_depths else 0
            del_med = f"{del_med_val:.3f}"
        except Exception:
            del_med = ""
            del_med_val = None

        ratio = ""
        if flank_med_val is not None and flank_med_val > 0 and del_med_val is not None:
            ratio = f"{(del_med_val / flank_med_val):.6f}"

        callable_flag = "FALSE"
        if flank_med_val is not None and flank_med_val >= ERM41_CALLABLE_MIN_FLANK_MED:
            callable_flag = "TRUE"

        out.write("\t".join([
            iso,
            str(erm_len),
            erm_med,
            left_med,
            right_med,
            flank_med,
            del_range_str,
            del_med,
            ratio,
            callable_flag,
        ]) + "\n")

# -----------------------
# erm41 coverage plots (per isolate)
# -----------------------
import matplotlib.pyplot as plt

PLOT_DIR = os.path.join(OUTDIR, "summary", "erm41_coverage_plots")
os.makedirs(PLOT_DIR, exist_ok=True)

ERM41_LEN = ERM41_END - ERM41_START + 1

def plot_erm41_coverage(isolate: str, depths: Dict[int, int]) -> None:
    """
    Plot per-base coverage across erm41 gene (gene coordinates),
    with shaded deletion regions only.
    """
    gene_pos = []
    gene_depth = []

    for i in range(ERM41_LEN):
        gpos = i + 1
        ref_pos = ERM41_START + i
        gene_pos.append(gpos)
        gene_depth.append(depths.get(ref_pos, 0))

    if not gene_depth:
        return

    max_depth = max(gene_depth)
    ymax = max_depth * 1.10 if max_depth > 0 else 10

    fig, ax = plt.subplots(figsize=(10, 4))

    # Coverage line
    ax.plot(gene_pos, gene_depth, linewidth=1)

    # Shaded regions
    ax.axvspan(64, 65, color="orange", alpha=0.25)
    ax.axvspan(159, 432, color="red", alpha=0.15)

    # Axis limits
    ax.set_xlim(0, ERM41_LEN)
    ax.set_ylim(0, ymax)

    # Axis labels
    ax.set_xlabel("erm41 gene position")
    ax.set_ylabel("Depth")

    # X ticks: only start and end
    ax.set_xticks([0, ERM41_LEN])

    ax.set_title(f"{isolate} – erm41 coverage")

    plt.tight_layout()

    out_png = os.path.join(PLOT_DIR, f"{isolate}.erm41_coverage.png")
    plt.savefig(out_png, dpi=150)
    plt.close(fig)


# Generate plots
for iso in isolates:
    depth_tsv = os.path.join(OUTDIR, iso, "variants", "targets_depth.tsv")
    depths = read_depth_tsv(depth_tsv, REF_CONTIG)
    if depths:
        plot_erm41_coverage(iso, depths)

# -----------------------
# Write combined Excel summary (3 tabs) - robust TSV streaming
# -----------------------
excel_out = os.path.join(OUTDIR, "summary", "myco_prediction_summary.xlsx")

def tsv_to_sheet(ws, tsv_path: str) -> None:
    """
    Write a TSV file to an openpyxl worksheet.
    Uses the header column count to pad/truncate inconsistent rows.
    """
    if (not os.path.exists(tsv_path)) or os.path.getsize(tsv_path) == 0:
        ws.append(["(empty)"])
        return

    with open(tsv_path, "r") as f:
        header = f.readline().rstrip("\n")
        if not header:
            ws.append(["(empty)"])
            return

        cols = header.split("\t")
        ncol = len(cols)
        ws.append(cols)

        for line in f:
            line = line.rstrip("\n")
            if not line:
                continue
            parts = line.split("\t")

            # pad/truncate to match header width
            if len(parts) < ncol:
                parts += [""] * (ncol - len(parts))
            elif len(parts) > ncol:
                # Keep extra fields glued into the last column so nothing is lost
                parts = parts[:ncol-1] + ["\t".join(parts[ncol-1:])]

            ws.append(parts)

try:
    from openpyxl import Workbook

    wb = Workbook()

    # variants sheet
    ws1 = wb.active
    ws1.title = "variants"
    tsv_to_sheet(ws1, sites_out)

    # truncation sheet
    ws2 = wb.create_sheet("truncation")
    tsv_to_sheet(ws2, trunc_out)

    # blast sheet
    ws3 = wb.create_sheet("blast")
    tsv_to_sheet(ws3, blast_out)

    wb.save(excel_out)
    print(f"Wrote: {excel_out}")

except Exception as e:
    print(f"WARNING: Could not write Excel summary ({excel_out}): {e}")


# -----------------------
# Status concatenation (Step1 + Step2)
# -----------------------
concat_status_files(OUTDIR, isolates)

print(f"Wrote: {blast_out}")
print(f"Wrote: {sites_out}")
print(f"Wrote: {trunc_out}")
print(f"Wrote: {os.path.join(OUTDIR, 'status', 'status_step1_all.tsv')}")
print(f"Wrote: {os.path.join(OUTDIR, 'status', 'status_step2_all.tsv')}")

